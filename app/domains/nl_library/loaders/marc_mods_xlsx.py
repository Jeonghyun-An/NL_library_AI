"""
marc_mods_xlsx.py — 국립중앙도서관 도서 메타 엑셀/CSV 로더 (CatalogLoader)

컬럼: CONTENTS_ID | TITLE_INFO | MARC | MODS
  MARC 있음 → marc_parser, MARC 없음 + MODS 있음 → mods_parser, 둘 다 없음 → title만
"""
import csv
import logging
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook

from domains.base import ParsedRecord, split_core_extra
from domains.nl_library.loaders.marc_parser import parse as parse_marc
from domains.nl_library.loaders.mods_parser import parse as parse_mods

log = logging.getLogger(__name__)


def _read_rows(file_path: str) -> tuple[list[str], list[list[str]]]:
    """
    xlsx 또는 csv 파일 읽기 → (헤더, 행 리스트) 반환
    xlsx 실패 시 csv로 fallback
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"파일 없음: {file_path}")

    # 1차: xlsx 시도
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        all_rows = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append([str(cell).strip() if cell is not None else "" for cell in row])
        wb.close()

        if not all_rows:
            raise ValueError("빈 엑셀")

        headers = [h.upper() for h in all_rows[0]]
        data_rows = all_rows[1:]
        log.info(f"xlsx 로드 성공: {len(data_rows)}행")
        return headers, data_rows
    except Exception as xlsx_err:
        log.warning(f"xlsx 로드 실패 ({xlsx_err}), csv로 시도")

    # 2차: csv fallback (다양한 인코딩 + 구분자 시도)
    for encoding in ["utf-8-sig", "utf-8", "cp949", "euc-kr"]:
        for delimiter in [",", "\t", "|"]:
            try:
                with open(path, encoding=encoding, newline="") as f:
                    reader = csv.reader(f, delimiter=delimiter)
                    all_rows = [[cell.strip() for cell in row] for row in reader if row]

                    if not all_rows or len(all_rows[0]) < 2:
                        continue

                    headers = [h.upper() for h in all_rows[0]]
                    data_rows = all_rows[1:]
                    log.info(f"csv 로드 성공: encoding={encoding}, delimiter={repr(delimiter)}, {len(data_rows)}행")
                    return headers, data_rows
            except Exception:
                continue

    raise ValueError(f"xlsx/csv 모두 로드 실패: {file_path}")


def load_xlsx(file_path: str) -> list[dict]:
    """엑셀/CSV 파일 읽기 → 파싱된 레코드(dict) 리스트 반환 (레거시 API)."""
    headers, data_rows = _read_rows(file_path)
    log.info(f"헤더: {headers}")

    try:
        idx_cnts = headers.index("CONTENTS_ID")
    except ValueError:
        raise ValueError(f"CONTENTS_ID 컬럼 없음. 헤더: {headers}")

    idx_title = headers.index("TITLE_INFO") if "TITLE_INFO" in headers else None
    idx_marc = headers.index("MARC") if "MARC" in headers else None
    idx_mods = headers.index("MODS") if "MODS" in headers else None

    records = []
    for row_num, row in enumerate(data_rows, start=2):
        if len(row) <= idx_cnts:
            continue

        cnts_id = row[idx_cnts].strip()
        if not cnts_id:
            continue

        title_raw = row[idx_title].strip() if idx_title is not None and len(row) > idx_title else ""
        marc_raw = row[idx_marc].strip() if idx_marc is not None and len(row) > idx_marc else ""
        mods_raw = row[idx_mods].strip() if idx_mods is not None and len(row) > idx_mods else ""

        try:
            marc_parsed = parse_marc(marc_raw) if marc_raw else {}
            mods_parsed = parse_mods(mods_raw) if mods_raw else {}

            if marc_parsed or mods_parsed:
                # MARC 우선, 빈 필드는 MODS로 보완
                parsed = {**mods_parsed, **{k: v for k, v in marc_parsed.items() if v is not None}}
                parsed["title"] = parsed.get("title") or title_raw
                if marc_parsed and mods_parsed:
                    parsed["source_format"] = "MARC+MODS"
                elif marc_parsed:
                    parsed["source_format"] = "MARC"
                else:
                    parsed["source_format"] = "MODS"
            else:
                parsed = {
                    "title": title_raw or cnts_id,
                    "source_format": "NONE",
                }

            parsed["cnts_id"] = cnts_id
            records.append(parsed)

        except Exception as e:
            log.error(f"[row {row_num}][{cnts_id}] 파싱 실패: {e}")
            # 파싱 실패해도 title이라도 저장
            records.append({
                "cnts_id": cnts_id,
                "title": title_raw or cnts_id,
                "source_format": "ERROR",
            })

    marc_count  = sum(1 for r in records if r.get("source_format") == "MARC")
    mods_count  = sum(1 for r in records if r.get("source_format") == "MODS")
    both_count  = sum(1 for r in records if r.get("source_format") == "MARC+MODS")
    other_count = len(records) - marc_count - mods_count - both_count
    log.info(f"총 {len(records)}건 로드 (MARC: {marc_count}, MODS: {mods_count}, MARC+MODS: {both_count}, 기타: {other_count})")

    return records


class MarcModsXlsxLoader:
    """CatalogLoader 구현 — MARC/MODS 엑셀을 ParsedRecord 스트림으로."""

    loader_id = "nl_marc_mods_xlsx"

    def detect(self, file_path: str, headers: list[str]) -> bool:
        return "CONTENTS_ID" in [h.upper() for h in headers]

    def load(self, file_path: str) -> Iterator[ParsedRecord]:
        for rec in load_xlsx(file_path):
            cnts_id = rec.pop("cnts_id")
            source_format = rec.get("source_format", "")
            core, extra = split_core_extra(rec)
            core["cnts_id"] = cnts_id
            yield ParsedRecord(
                source_id=cnts_id,
                core=core,
                extra=extra,
                source_format=source_format,
            )
