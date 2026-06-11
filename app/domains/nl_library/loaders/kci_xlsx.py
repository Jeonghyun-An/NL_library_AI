"""
kci_xlsx.py — KCI 논문 메타데이터 xlsx 로더 (CatalogLoader)

컬럼: KCI_FI_ID | ARTI_ID | GRADE | TITLE_KR | TITLE_EN |
      AUTHORS | INSTITUTION | JOURNAL | VOL_ISSUE | PAGES |
      YEAR_MONTH | KCI_CITATIONS | WOS_CITATIONS
"""
import logging
from typing import Iterator

from openpyxl import load_workbook

from domains.base import ParsedRecord, split_core_extra

log = logging.getLogger(__name__)

# 레거시 컬럼 매핑 — 기존 library_catalog 컬럼 재사용 (동작 불변)
_COL_MAP = {
    "KCI_FI_ID":      "cnts_id",
    "ARTI_ID":        "record_id",
    "GRADE":          "grade",
    "TITLE_KR":       "title",
    "TITLE_EN":       "title_remainder",
    "AUTHORS":        "personal_author",
    "INSTITUTION":    "corporate_author",
    "JOURNAL":        "series_title",
    "VOL_ISSUE":      "vol_issue",
    "PAGES":          "extent",
    "YEAR_MONTH":     "pub_date",
    "KCI_CITATIONS":  "kci_citations",
    "WOS_CITATIONS":  "wos_citations",
}

# 의미가 어긋난 레거시 매핑의 본래 이름 — extra JSONB에 함께 기록 (장기 이행용)
_TRUE_NAMES = {
    "TITLE_EN": "title_en",
    "JOURNAL":  "journal",
    "PAGES":    "pages",
}


def load_kci_xlsx(file_path: str) -> list[dict]:
    """KCI xlsx → 파싱된 레코드(dict) 리스트 (레거시 API)."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append([str(c).strip() if c is not None else "" for c in row])
    wb.close()

    if not all_rows:
        raise ValueError("빈 엑셀 파일")

    headers = [h.upper() for h in all_rows[0]]
    log.info(f"KCI xlsx 헤더: {headers}")

    if "KCI_FI_ID" not in headers:
        raise ValueError(f"KCI_FI_ID 컬럼 없음. 헤더: {headers}")

    col_idx = {col: headers.index(col) for col in _COL_MAP if col in headers}

    records = []
    for row in all_rows[1:]:
        cnts_id_col = col_idx.get("KCI_FI_ID")
        if cnts_id_col is None or len(row) <= cnts_id_col:
            continue

        cnts_id = row[cnts_id_col].strip()
        if not cnts_id:
            continue

        rec: dict = {"source_format": "KCI", "genre": "paper", "language": "ko"}
        true_names: dict = {}
        for col, field in _COL_MAP.items():
            idx = col_idx.get(col)
            if idx is None or len(row) <= idx:
                continue
            val = row[idx].strip()
            if not val:
                continue

            # 숫자 필드
            if field in ("kci_citations", "wos_citations"):
                try:
                    rec[field] = int(val)
                except ValueError:
                    rec[field] = 0
            else:
                rec[field] = val

            if col in _TRUE_NAMES:
                true_names[_TRUE_NAMES[col]] = val

        if not rec.get("title"):
            rec["title"] = cnts_id
        rec["_true_names"] = true_names
        records.append(rec)

    log.info(f"KCI 메타데이터 {len(records)}건 파싱 완료")
    return records


class KciXlsxLoader:
    """CatalogLoader 구현 — KCI xlsx를 ParsedRecord 스트림으로."""

    loader_id = "kci_xlsx"

    def detect(self, file_path: str, headers: list[str]) -> bool:
        return "KCI_FI_ID" in [h.upper() for h in headers]

    def load(self, file_path: str) -> Iterator[ParsedRecord]:
        for rec in load_kci_xlsx(file_path):
            true_names = rec.pop("_true_names", {})
            cnts_id = rec.pop("cnts_id")
            core, extra = split_core_extra(rec)
            core["cnts_id"] = cnts_id
            core["doc_type"] = "paper"
            extra.update(true_names)
            yield ParsedRecord(
                source_id=cnts_id,
                core=core,
                extra=extra,
                source_format="KCI",
            )
