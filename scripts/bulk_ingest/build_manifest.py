"""
build_manifest.py — 로컬 PC에서 실행하는 메타↔PDF 매칭/검증 도구

엑셀 메타데이터와 PDF 디렉토리를 스캔하여:
  1. PDF 파일명(stem) ↔ 메타 ID 매칭 (정규화: strip + 끝 '_' 제거)
  2. manifest.jsonl 생성 — 1행 = {book_id, file, object_key, size, title}
  3. validation_report.json 생성 — matched / pdf_without_meta / meta_without_pdf /
                                    duplicates / zero_size (+ --deep-check 시 unreadable_pdf)

object_key 는 originals/{book_id}/{book_id}.pdf 구조 — 썸네일 폴백이 이 prefix 를
스캔하므로 플랫 키(originals/{id}.pdf)는 사용하지 않는다.

표준 라이브러리 + openpyxl 만 의존 (앱 코드 임포트 없음 — 로컬에서 독립 실행).

사용:
  python build_manifest.py --excel meta.xlsx --pdf-dir ./pdfs --out ./out
  python build_manifest.py --excel a.xlsx b.xlsx --pdf-dir ./pdfs --out ./out \
         --match-by arti_id --deep-check
"""
import argparse
import json
import os
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


# 매칭 기준 컬럼 (대문자 비교)
_MATCH_COLUMNS = {
    "kci_fi_id": "KCI_FI_ID",
    "arti_id": "ARTI_ID",
    "contents_id": "CONTENTS_ID",
}
_TITLE_COLUMNS = ("TITLE_KR", "TITLE_INFO", "TITLE")


def normalize_book_id(raw: str) -> str:
    """앱의 _normalize_book_id 와 동일 — strip + 끝 '_' 제거."""
    return raw.strip().rstrip("_").strip()


def _read_meta(meta_path: str, match_col: str) -> dict[str, dict]:
    """엑셀/CSV에서 {정규화 ID: {raw_id, title}} 반환. 확장자로 분기 (.csv는 스트리밍)."""
    if meta_path.lower().endswith(".csv"):
        return _read_meta_csv(meta_path, match_col)
    return _read_meta_xlsx(meta_path, match_col)


def _read_meta_csv(csv_path: str, match_col: str) -> dict[str, dict]:
    """CSV 스트리밍 파싱 — 대용량(수십만 행)에 xlsx보다 빠르고 가볍다."""
    import csv as _csv

    for encoding in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            out: dict[str, dict] = {}
            with open(csv_path, encoding=encoding, newline="") as f:
                reader = _csv.reader(f)
                header = next(reader, None)
                if not header:
                    return {}
                cols = {h.strip().upper(): i for i, h in enumerate(header)}
                if match_col not in cols:
                    raise ValueError(f"매칭 컬럼 '{match_col}' 없음. 헤더: {list(cols)}")
                id_idx = cols[match_col]
                title_idx = next((cols[c] for c in _TITLE_COLUMNS if c in cols), None)
                for row in reader:
                    if id_idx >= len(row):
                        continue
                    raw_id = (row[id_idx] or "").strip()
                    if not raw_id:
                        continue
                    norm = normalize_book_id(raw_id)
                    title = ""
                    if title_idx is not None and title_idx < len(row):
                        title = (row[title_idx] or "").strip()
                    out[norm] = {"raw_id": raw_id, "title": title}
            return out
        except UnicodeDecodeError:
            continue
    raise ValueError(f"CSV 인코딩 판별 실패: {csv_path}")


def _read_meta_xlsx(excel_path: str, match_col: str) -> dict[str, dict]:
    """엑셀에서 {정규화 ID: {raw_id, title}} 반환."""
    if load_workbook is None:
        raise RuntimeError("openpyxl 미설치 — pip install openpyxl")
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        wb.close()
        return {}
    cols = {str(h).strip().upper(): i for i, h in enumerate(header) if h is not None}
    if match_col not in cols:
        wb.close()
        raise ValueError(f"매칭 컬럼 '{match_col}' 없음. 헤더: {list(cols)}")
    id_idx = cols[match_col]
    title_idx = next((cols[c] for c in _TITLE_COLUMNS if c in cols), None)

    out: dict[str, dict] = {}
    for row in rows:
        if id_idx >= len(row) or row[id_idx] is None:
            continue
        raw_id = str(row[id_idx]).strip()
        if not raw_id:
            continue
        norm = normalize_book_id(raw_id)
        title = ""
        if title_idx is not None and title_idx < len(row) and row[title_idx] is not None:
            title = str(row[title_idx]).strip()
        out[norm] = {"raw_id": raw_id, "title": title}
    wb.close()
    return out


def _scan_pdfs(pdf_dir: str) -> tuple[dict[str, Path], dict[str, list[str]]]:
    """PDF 디렉토리 재귀 스캔 → ({정규화 ID: Path}, {중복 ID: [경로...]})."""
    by_id: dict[str, Path] = {}
    duplicates: dict[str, list[str]] = {}
    root = Path(pdf_dir)
    for path in sorted(root.rglob("*.pdf")):
        if not path.is_file():
            continue
        norm = normalize_book_id(path.stem)
        if norm in by_id:
            duplicates.setdefault(norm, [str(by_id[norm])]).append(str(path))
            continue  # 첫 번째 발견본을 유지
        by_id[norm] = path
    return by_id, duplicates


def _is_readable_pdf(path: Path) -> bool:
    """헤더 매직바이트(%PDF)로 간이 검사 (--deep-check 없이도 0바이트 외 손상 일부 감지)."""
    try:
        with open(path, "rb") as f:
            return f.read(5).startswith(b"%PDF")
    except OSError:
        return False


def build_manifest(
    excel_paths: list[str],
    pdf_dir: str,
    out_dir: str,
    match_by: str = "kci_fi_id",
    deep_check: bool = False,
    limit: int | None = None,
) -> dict:
    match_col = _MATCH_COLUMNS.get(match_by, match_by.upper())

    meta: dict[str, dict] = {}
    for xp in excel_paths:
        meta.update(_read_meta(xp, match_col))

    pdfs_by_id, duplicates = _scan_pdfs(pdf_dir)

    meta_ids = set(meta)
    pdf_ids = set(pdfs_by_id)

    matched_rows: list[dict] = []
    zero_size: list[str] = []
    unreadable: list[str] = []

    for book_id in sorted(meta_ids & pdf_ids):
        path = pdfs_by_id[book_id]
        size = path.stat().st_size
        if size == 0:
            zero_size.append(book_id)
            continue
        if deep_check and not _is_readable_pdf(path):
            unreadable.append(book_id)
            continue
        matched_rows.append({
            "book_id": book_id,
            "file": str(path),
            "object_key": f"originals/{book_id}/{book_id}.pdf",
            "size": size,
            "title": meta[book_id]["title"],
        })

    # --limit: 파일럿 서브셋 (리포트는 전체 기준, 매니페스트만 N건)
    written_rows = matched_rows[:limit] if limit else matched_rows

    report = {
        "matched": len(matched_rows),
        "manifest_written": len(written_rows),
        "total_meta": len(meta_ids),
        "total_pdf": len(pdf_ids),
        "meta_without_pdf": sorted(meta_ids - pdf_ids),
        "pdf_without_meta": sorted(pdf_ids - meta_ids),
        "duplicates": duplicates,
        "zero_size": sorted(zero_size),
        "unreadable_pdf": sorted(unreadable),
    }

    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    with open(out / "manifest.jsonl", "w", encoding="utf-8") as f:
        for row in written_rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")
    with open(out / "validation_report.json", "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    return report


def _print_summary(report: dict) -> None:
    print("── 매니페스트 검증 결과 ─────────────────────")
    print(f"  매칭 성공      : {report['matched']}")
    print(f"  매니페스트 기록 : {report['manifest_written']}")
    print(f"  메타 총계      : {report['total_meta']}")
    print(f"  PDF 총계       : {report['total_pdf']}")
    print(f"  메타만 (PDF 없음): {len(report['meta_without_pdf'])}")
    print(f"  PDF만 (메타 없음): {len(report['pdf_without_meta'])}")
    print(f"  중복 ID        : {len(report['duplicates'])}")
    print(f"  0바이트        : {len(report['zero_size'])}")
    if report["unreadable_pdf"]:
        print(f"  손상 PDF       : {len(report['unreadable_pdf'])}")


def main() -> None:
    ap = argparse.ArgumentParser(description="메타↔PDF 매칭 매니페스트 생성")
    ap.add_argument("--excel", nargs="+", required=True, help="메타 파일(들) — .xlsx 또는 .csv")
    ap.add_argument("--pdf-dir", required=True, help="PDF 루트 디렉토리 (재귀 스캔)")
    ap.add_argument("--out", required=True, help="manifest.jsonl / validation_report.json 출력 디렉토리")
    ap.add_argument("--match-by", default="kci_fi_id",
                    choices=list(_MATCH_COLUMNS), help="매칭 기준 컬럼")
    ap.add_argument("--deep-check", action="store_true",
                    help="PDF 헤더 매직바이트 검사 (손상 파일 감지)")
    ap.add_argument("--limit", type=int, default=None,
                    help="파일럿용 — 매칭된 것 중 앞 N건만 매니페스트에 기록 (리포트는 전체)")
    args = ap.parse_args()

    report = build_manifest(
        excel_paths=args.excel,
        pdf_dir=args.pdf_dir,
        out_dir=args.out,
        match_by=args.match_by,
        deep_check=args.deep_check,
        limit=args.limit,
    )
    _print_summary(report)
    print(f"\n→ {os.path.join(args.out, 'manifest.jsonl')}")
    print(f"→ {os.path.join(args.out, 'validation_report.json')}")


if __name__ == "__main__":
    main()
