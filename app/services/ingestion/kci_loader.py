"""
kci_loader.py — KCI 논문 메타데이터 xlsx 파싱

컬럼: kci_fi_id | arti_id | grade | title_kr | title_en |
      authors | institution | journal | vol_issue | pages |
      year_month | kci_citations | wos_citations
"""
import logging
from openpyxl import load_workbook

log = logging.getLogger(__name__)

_COL_MAP = {
    "KCI_FI_ID":      "cnts_id",
    "ARTI_ID":        "record_id",
    "GRADE":          "grade",
    "TITLE_KR":       "title",
    "TITLE_EN":       "title_remainder",
    "AUTHORS":        "personal_author",
    "INSTITUTION":    "corporate_author",
    "JOURNAL":        "series_title",
    "VOL_ISSUE":      "vol_issue",
    "PAGES":          "extent",
    "YEAR_MONTH":     "pub_date",
    "KCI_CITATIONS":  "kci_citations",
    "WOS_CITATIONS":  "wos_citations",
}


def load_kci_xlsx(file_path: str) -> list[dict]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append([str(c).strip() if c is not None else "" for c in row])
    wb.close()

    if not all_rows:
        raise ValueError("빈 엑셀 파일")

    headers = [h.upper() for h in all_rows[0]]
    log.info(f"KCI xlsx 헤더: {headers}")

    if "KCI_FI_ID" not in headers:
        raise ValueError(f"KCI_FI_ID 컬럼 없음. 헤더: {headers}")

    col_idx = {col: headers.index(col) for col in _COL_MAP if col in headers}

    records = []
    for row_num, row in enumerate(all_rows[1:], start=2):
        cnts_id_col = col_idx.get("KCI_FI_ID")
        if cnts_id_col is None or len(row) <= cnts_id_col:
            continue

        cnts_id = row[cnts_id_col].strip()
        if not cnts_id:
            continue

        rec: dict = {"source_format": "KCI", "genre": "paper", "language": "ko"}
        for col, field in _COL_MAP.items():
            idx = col_idx.get(col)
            if idx is None or len(row) <= idx:
                continue
            val = row[idx].strip()
            if not val:
                continue

            # 숫자 필드
            if field in ("kci_citations", "wos_citations"):
                try:
                    rec[field] = int(val)
                except ValueError:
                    rec[field] = 0
            else:
                rec[field] = val

        # title 필수
        if not rec.get("title"):
            rec["title"] = cnts_id

        records.append(rec)

    log.info(f"KCI 메타데이터 {len(records)}건 파싱 완료")
    return records
