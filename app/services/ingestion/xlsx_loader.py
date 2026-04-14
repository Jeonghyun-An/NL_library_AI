"""
xlsx_loader.py
엑셀/CSV 로더

컬럼: CONTENTS_ID | TITLE_INFO | MARC | MODS

MARC 있음 → marc_parser
MARC 없음, MODS 있음 → mods_parser
둘 다 없음 → title만 저장
"""
import csv
import logging
from pathlib import Path
from openpyxl import load_workbook

from services.ingestion.marc_parser import parse as parse_marc
from services.ingestion.mods_parser import parse as parse_mods

log = logging.getLogger(__name__)


def _read_rows(file_path: str) -> tuple[list[str], list[list[str]]]:
    """
    xlsx 또는 csv 파일 읽기 → (헤더, 행 리스트) 반환
    xlsx 실패 시 csv로 fallback
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"파일 없음: {file_path}")

    # 1차: xlsx 시도
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        all_rows = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append([str(cell).strip() if cell is not None else "" for cell in row])
        wb.close()

        if not all_rows:
            raise ValueError("빈 엑셀")

        headers = [h.upper() for h in all_rows[0]]
        data_rows = all_rows[1:]
        log.info(f"xlsx 로드 성공: {len(data_rows)}행")
        return headers, data_rows
    except Exception as xlsx_err:
        log.warning(f"xlsx 로드 실패 ({xlsx_err}), csv로 시도")

    # 2차: csv fallback (다양한 인코딩 + 구분자 시도)
    for encoding in ["utf-8-sig", "utf-8", "cp949", "euc-kr"]:
        for delimiter in [",", "\t", "|"]:
            try:
                with open(path, encoding=encoding, newline="") as f:
                    # 첫 줄 읽어서 구분자 판별
                    sample = f.read(4096)
                    f.seek(0)

                    reader = csv.reader(f, delimiter=delimiter)
                    all_rows = [[cell.strip() for cell in row] for row in reader if row]

                    if not all_rows or len(all_rows[0]) < 2:
                        continue

                    headers = [h.upper() for h in all_rows[0]]
                    data_rows = all_rows[1:]
                    log.info(f"csv 로드 성공: encoding={encoding}, delimiter={repr(delimiter)}, {len(data_rows)}행")
                    return headers, data_rows
            except Exception:
                continue

    raise ValueError(f"xlsx/csv 모두 로드 실패: {file_path}")


def load_xlsx(file_path: str) -> list[dict]:
    """
    엑셀/CSV 파일 읽기 → 파싱된 레코드 리스트 반환
    """
    headers, data_rows = _read_rows(file_path)
    log.info(f"헤더: {headers}")

    # 컬럼 인덱스 찾기
    try:
        idx_cnts = headers.index("CONTENTS_ID")
    except ValueError:
        raise ValueError(f"CONTENTS_ID 컬럼 없음. 헤더: {headers}")

    idx_title = headers.index("TITLE_INFO") if "TITLE_INFO" in headers else None
    idx_marc = headers.index("MARC") if "MARC" in headers else None
    idx_mods = headers.index("MODS") if "MODS" in headers else None

    records = []
    for row_num, row in enumerate(data_rows, start=2):
        # 컬럼 수가 부족한 행 스킵
        if len(row) <= idx_cnts:
            continue

        cnts_id = row[idx_cnts].strip()
        if not cnts_id:
            continue

        title_raw = row[idx_title].strip() if idx_title is not None and len(row) > idx_title else ""
        marc_raw = row[idx_marc].strip() if idx_marc is not None and len(row) > idx_marc else ""
        mods_raw = row[idx_mods].strip() if idx_mods is not None and len(row) > idx_mods else ""

        try:
            if marc_raw:
                parsed = parse_marc(marc_raw)
                parsed["title"] = parsed.get("title") or title_raw
                parsed["source_format"] = "MARC"
            elif mods_raw:
                parsed = parse_mods(mods_raw)
                parsed["title"] = parsed.get("title") or title_raw
                parsed["source_format"] = "MODS"
            else:
                parsed = {
                    "title": title_raw or cnts_id,
                    "source_format": "NONE",
                }

            parsed["cnts_id"] = cnts_id
            records.append(parsed)

        except Exception as e:
            log.error(f"[row {row_num}][{cnts_id}] 파싱 실패: {e}")
            # 파싱 실패해도 title이라도 저장
            records.append({
                "cnts_id": cnts_id,
                "title": title_raw or cnts_id,
                "source_format": "ERROR",
            })

    marc_count = sum(1 for r in records if r.get("source_format") == "MARC")
    mods_count = sum(1 for r in records if r.get("source_format") == "MODS")
    log.info(f"총 {len(records)}건 로드 (MARC: {marc_count}, MODS: {mods_count}, 기타: {len(records) - marc_count - mods_count})")

    return records