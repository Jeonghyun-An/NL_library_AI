"""nl_library 파서 플러그인 — CatalogLoader 구현체 레지스트리."""
import csv
import logging
from pathlib import Path

log = logging.getLogger(__name__)


def build_loaders():
    from domains.nl_library.loaders.kci_xlsx import KciXlsxLoader
    from domains.nl_library.loaders.marc_mods_xlsx import MarcModsXlsxLoader

    return [KciXlsxLoader(), MarcModsXlsxLoader()]


def read_headers(file_path: str) -> list[str]:
    """파일의 헤더 행만 읽기 (loader.detect() 용 — 전체 로드 없이)."""
    path = Path(file_path)
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True, max_row=1):
            wb.close()
            return [str(c).strip().upper() if c is not None else "" for c in row]
        wb.close()
        return []
    except Exception:
        pass

    for encoding in ["utf-8-sig", "utf-8", "cp949", "euc-kr"]:
        for delimiter in [",", "\t", "|"]:
            try:
                with open(path, encoding=encoding, newline="") as f:
                    row = next(csv.reader(f, delimiter=delimiter), None)
                    if row and len(row) >= 2:
                        return [c.strip().upper() for c in row]
            except Exception:
                continue
    return []


def resolve_loader(file_path: str, loader_id: str | None = None):
    """loader_id 지정 시 직접 매칭, 아니면 헤더 기반 자동 판별. 없으면 None."""
    loaders = build_loaders()
    if loader_id:
        for ld in loaders:
            if ld.loader_id == loader_id:
                return ld
        return None
    headers = read_headers(file_path)
    for ld in loaders:
        if ld.detect(file_path, headers):
            return ld
    return None
